from celery import shared_task, current_task
from .models import Proyeccion
from django.core.mail import send_mail as _send_mail
from django.core.mail import EmailMessage
from django.utils import timezone
from openpyxl import Workbook
from io import BytesIO

TODAY = timezone.now().date()

@shared_task
def send_mail_proyeccion():
    proyeciones = Proyeccion.objects.select_related('estacion', 'parte').all()
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'id'
    ws['B1'] = 'estacion'
    ws['C1'] = 'proyecto'
    ws['D1'] = 'escenario'
    ws['E1'] = 'banda'
    ws['F1'] = 'agrupadores'
    ws['G1'] = 'rfe'
    ws['H1'] = 'parte'
    ws['I1'] = 'estado_proyeccion'
    ws['J1'] = 'cantidad_estimada'
    count = 2
    for proyeccion in proyeciones:
        ws.cell(row=count,column=1).value = proyeccion.id
        ws.cell(row=count,column=2).value = proyeccion.estacion.site_name
        ws.cell(row=count,column=3).value = proyeccion.proyecto
        ws.cell(row=count,column=4).value = proyeccion.escenario
        ws.cell(row=count,column=5).value = proyeccion.banda
        ws.cell(row=count,column=6).value = proyeccion.agrupadores
        ws.cell(row=count,column=7).value = proyeccion.rfe
        ws.cell(row=count,column=8).value = proyeccion.parte.parte_nokia
        ws.cell(row=count,column=9).value = proyeccion.estado_proyeccion
        ws.cell(row=count,column=10).value = proyeccion.cantidad_estimada
        count += 1
    wb.save(output)
    filename = 'Proyeccion-' + TODAY.strftime('%Y-%m-%d') + '.xlsx'
    content = output.getvalue()
    mimetype = 'application/vnd.ms-excel'
    message = EmailMessage(
        'Proyecciones de Hardware',
        'Proyecciones de Hardware ' + TODAY.strftime('%Y-%m-%d'),
        'notificaciones@nihardware.com',
        [
        'jbri.gap@nokia.com',
        'hw.proyections@nokia.com',
        'administration.hw@nokia.com',
        'hw_control_2.ni@nokia.com',
        'hw_control.ni@nokia.com',
        'csp_support.ni_co@nokia.com',
        'camilo.lozano@nokia.com',
        ],
        )
    message.attach(filename, content, mimetype)
    message.send(fail_silently=False)
    return { 'status_code':200 }
